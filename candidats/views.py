# candidats/views.py
from rest_framework import status, viewsets
from rest_framework.decorators import api_view, permission_classes, parser_classes, action
from rest_framework.response import Response
from django.db.models import Count, Q, F, Avg, Sum
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q
from authentication.models import User
from django.utils import timezone
from datetime import date, timedelta
from candidats.utils.utils import create_notification
from candidats.utils.pdf_generator import generer_fiche_enrollement
from django.core.mail import EmailMessage
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from .permissions import IsAdminAcademique
import qrcode
from io import BytesIO
import base64
from .utils.pdf_generator import generer_fiche_enrollement
from .models import Candidat, Dossier, Document,  Notification
from .serializers import (
    CandidatEnrollementSerializer,
    CandidatListSerializer,
    CandidatDetailSerializer,
    DossierValidationSerializer
)
from .permissions import IsResponsableFiliere, IsAdminAcademique
from authentication.models import CodeQuitus
from configurations.models import Filiere
def send_validation_email_async(candidat_id):
    """Envoyer l'email de validation en arrière-plan (NON BLOQUANT)"""
    try:
        # Importer DANS le thread pour éviter les problèmes de connexion DB
        from candidats.models import Candidat
        from candidats.utils.pdf_generator import generer_fiche_enrollement
        from django.template.loader import render_to_string
        from django.core.mail import EmailMessage
        from django.conf import settings
        import socket
        
        print(f"\n🧵 [THREAD EMAIL] Démarré pour candidat ID={candidat_id}")
        
        # Récupérer le candidat
        candidat = Candidat.objects.get(id=candidat_id)
        print(f"   ✅ Candidat: {candidat.matricule}")
        
        # Générer le PDF
        print("   📄 Génération PDF...")
        pdf_buffer = generer_fiche_enrollement(candidat)
        print(f"   ✅ PDF: {len(pdf_buffer.getvalue())} octets")
        
        # Préparer le contexte
        context = {
            'candidat': candidat,
            'filiere': candidat.filiere
        }
        
        # Rendre le template HTML
        print("   📧 Rendu template...")
        html_message = render_to_string(
            'emails/validation_enrollement.html',
            context
        )
        print(f"   ✅ Template: {len(html_message)} caractères")
        
        # Créer l'email
        print("   📧 Création EmailMessage...")
        email = EmailMessage(
            subject=f'✅ Validation enrôlement - {candidat.filiere.libelle}',
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[candidat.email],
        )
        email.content_subtype = "html"
        
        # Attacher le PDF
        filename = f'Fiche_Enrollement_{candidat.matricule}.pdf'
        email.attach(filename, pdf_buffer.getvalue(), 'application/pdf')
        print(f"   ✅ PDF attaché")
        
        # 🔥 ENVOYER AVEC TIMEOUT
        print("   📤 Envoi email (timeout 15s)...")
        socket.setdefaulttimeout(15)  # Timeout de 15 secondes
        
        result = email.send(fail_silently=False)
        
        if result == 1:
            print(f"   ✅ ✅ ✅ EMAIL ENVOYÉ avec succès à {candidat.email}")
        else:
            print(f"   ⚠️ Résultat: {result}")
        
    except socket.timeout:
        print(f"   ❌ TIMEOUT lors de l'envoi email (serveur SMTP ne répond pas)")
    except Exception as e:
        print(f"   ❌ Erreur email thread: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print(f"🧵 [THREAD EMAIL] Terminé\n")

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_enrollment_status(request):
    """Vérifier si le candidat est déjà enrôlé"""
    try:
        user = request.user
        print(f"\n{'='*60}")
        print(f"🔍 CHECK ENROLLMENT STATUS")
        print(f"{'='*60}")
        print(f"👤 User: {user.email} (role: {user.role})")
        
        # Vérifier si c'est un candidat
        if user.role != 'candidat':
            print(f"❌ Pas un candidat (role: {user.role})")
            return Response({
                'is_enrolled': False,
                'message': 'Utilisateur non candidat'
            })
        
        # Chercher le candidat
        try:
            candidat = Candidat.objects.get(user=user)
            print(f"✅ Candidat trouvé: {candidat.matricule}")
            print(f"   Statut: {candidat.statut_dossier}")
            print(f"   Nom: {candidat.prenom} {candidat.nom}")
            
            # Un candidat est considéré comme "enrôlé" si son dossier existe
            # peut être en_attente, complet, valide ou même rejete
            is_enrolled = candidat.statut_dossier in ['en_attente', 'complet', 'valide', 'rejete']
            
            print(f"📊 Is enrolled: {is_enrolled}")
            print(f"{'='*60}\n")
            
            return Response({
                'is_enrolled': is_enrolled,
                'statut_dossier': candidat.statut_dossier,
                'matricule': candidat.matricule,
                'nom_complet': f"{candidat.prenom} {candidat.nom}",
                'filiere': candidat.filiere.libelle if candidat.filiere else None,
                'photo_url': request.build_absolute_uri(
                    settings.MEDIA_URL + candidat.photo_path
                ) if candidat.photo_path else None
            })
            
        except Candidat.DoesNotExist:
            print(f"❌ Aucun candidat trouvé pour user {user.id}")
            print(f"{'='*60}\n")
            return Response({
                'is_enrolled': False,
                'message': 'Aucun profil candidat trouvé'
            })
            
    except Exception as e:
        print(f"❌ Erreur check enrollment: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'is_enrolled': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def mon_profil_view(request):
    """
    GET: Récupérer les informations du profil candidat
    PUT: Mettre à jour le profil (si nécessaire plus tard)
    """
    try:
        candidat = request.user.candidat
    except Candidat.DoesNotExist:
        return Response({
            'error': 'Profil candidat non trouvé'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        # Récupérer le code quitus associé
        code_quitus = ''
        try:
            quitus = CodeQuitus.objects.get(utilisateur=request.user)
            code_quitus = quitus.code
        except CodeQuitus.DoesNotExist:
            pass
        
        data = {
            'candidat': {
                'id': candidat.id,
                'matricule': candidat.matricule,
                'nom': candidat.nom,
                'prenom': candidat.prenom,
                'email': candidat.email,
                'telephone': candidat.telephone,
                'date_naissance': candidat.date_naissance,
                'lieu_naissance': candidat.lieu_naissance,
                'sexe': candidat.sexe,
                'code_quitus': code_quitus,  # ✅ Code quitus ajouté
                'statut_dossier': candidat.statut_dossier,
                'created_at': candidat.created_at
            }
        }
        return Response(data, status=status.HTTP_200_OK)
    
    elif request.method == 'PUT':
        # Pour une future mise à jour du profil (optionnel)
        return Response({
            'message': 'Mise à jour du profil non implémentée'
        }, status=status.HTTP_501_NOT_IMPLEMENTED)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def enrollement_view(request):
    """Compléter profil candidat - USER AUTHENTIFIÉ"""
    
    user = request.user
    print(f"\n{'='*80}")
    print(f"👤 USER AUTHENTIFIÉ: {user.email} (role: {user.role})")
    print(f"{'='*80}")
    
    # ✅ 1. VÉRIFIER RÔLE
    if user.role != 'candidat':
        print("❌ User n'est pas candidat")
        return Response({
            'error': 'Accès refusé',
            'message': 'Seuls les candidats peuvent accéder à cette fonctionnalité.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # ✅ 2. LOG STATUT ACTUEL
    try:
        candidat_existant = Candidat.objects.get(user=user)
        print(f"📋 Candidat existant: {candidat_existant.matricule} | Statut: {candidat_existant.statut_dossier}")
    except Candidat.DoesNotExist:
        print("📋 Aucun candidat trouvé")
    
    # ✅ 3. DEBUG FICHIERS (CRITIQUE)
    print(f"\n📂 FICHIERS REÇUS ({len(request.FILES)}):")
    for key in request.FILES.keys():
        file = request.FILES[key]
        print(f"  ✅ {key}: {file.name} ({file.size} bytes)")
    
    print(f"\n📥 DONNÉES ({len(request.data)}):")
    for key, value in request.data.items():
        print(f"  📝 {key}: {value}")
    
    # 🔥 FIX CRITIQUE : CRÉER UN DICT STANDARD AU LIEU DE .copy()
    # request.data contient déjà les fichiers ET les données
    # Pas besoin de les combiner manuellement
    
    # ✅ 4. VALIDATION SERIALIZER DIRECTEMENT AVEC request.data
    serializer = CandidatEnrollementSerializer(data=request.data, context={'request': request})
    
    print(f"\n🔍 VALIDATION...")
    if serializer.is_valid():
        print("✅ VALIDATION OK")
        
        try:
            print(f"\n💾 SAUVEGARDE...")
            candidat = serializer.save()
            print(f"✅ SUCCÈS: {candidat.matricule} | Statut: {candidat.statut_dossier}")
            
            return Response({
                'success': True,
                'message': 'Enrôlement réussi ! Votre dossier est complet.',
                'candidat': {
                    'matricule': candidat.matricule,
                    'nom_complet': f'{candidat.nom} {candidat.prenom}',
                    'statut_dossier': candidat.statut_dossier,
                },
                'next_steps': [
                    'Votre dossier a été soumis avec succès',
                    'Consultez votre tableau de bord pour suivre l\'état'
                ]
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            print(f"❌ ERREUR SAUVEGARDE: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({
                'error': 'Erreur sauvegarde',
                'message': str(e),
                'support': 'support@estlc.cm'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # ❌ ERREURS
    print(f"❌ ERREURS: {serializer.errors}")
    return Response({
        'error': 'Données invalides',
        'details': serializer.errors,
        'message': 'Veuillez corriger les erreurs indiquées.'
    }, status=status.HTTP_400_BAD_REQUEST)
# candidats/views.py - ResponsableFiliereViewSet COMPLET ET CORRIGÉ

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mon_dossier_view(request):
    """
    Récupérer le dossier complet du candidat connecté
    """
    user = request.user
    
    # Vérifier que c'est un candidat
    if user.role != 'candidat':
        return Response({
            'error': 'Accès refusé',
            'message': 'Seuls les candidats peuvent accéder à cette page.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Récupérer le candidat
    try:
        candidat = Candidat.objects.select_related(
            'user',
            'region',
            'departement',
            'bac',
            'serie',
            'mention',
            'filiere',
            'niveau',
            'centre_examen',
            'centre_depot',
            'valide_par'
        ).get(user=user)
    except Candidat.DoesNotExist:
        return Response({
            'error': 'Profil candidat non trouvé',
            'message': 'Veuillez compléter votre enrôlement.'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Récupérer le dossier
    try:
        dossier = Dossier.objects.select_related(
            'annee_scolaire'
        ).get(candidat=candidat)
    except Dossier.DoesNotExist:
        dossier = None
    
    # Récupérer les documents (✅ Correction: created_at au lieu de date_upload)
    documents = Document.objects.filter(candidat=candidat).order_by('-created_at')
    
    # Construire les URLs complètes pour les documents
    from django.conf import settings
    documents_data = []
    for doc in documents:
        doc_url = f"{settings.MEDIA_URL}{doc.chemin_fichier}" if doc.chemin_fichier else None
        if doc_url and not doc_url.startswith('http'):
            doc_url = request.build_absolute_uri(doc_url)
        
        documents_data.append({
            'id': doc.id,
            'type_document': doc.type_document,
            'nom_fichier': doc.nom_fichier,
            'nom_original': doc.nom_original,
            'url': doc_url,
            'taille_fichier': doc.taille_fichier,
            'extension': doc.extension,
            'is_verified': doc.is_verified,
            'date_upload': doc.created_at,  # ✅ Utiliser created_at
            'date_verification': doc.verified_at  # ✅ Correction du nom
        })
    
    # Construire l'URL de la photo
    photo_url = None
    if candidat.photo_path:
        photo_url = f"{settings.MEDIA_URL}{candidat.photo_path}"
        if not photo_url.startswith('http'):
            photo_url = request.build_absolute_uri(photo_url)
    
    # Construire la réponse
    response_data = {
        'candidat': {
            'id': candidat.id,
            'matricule': candidat.matricule,
            'nom': candidat.nom,
            'prenom': candidat.prenom,
            'email': candidat.email,
            'telephone': candidat.telephone,
            'telephone_secondaire': candidat.telephone_secondaire,
            'date_naissance': candidat.date_naissance,
            'lieu_naissance': candidat.lieu_naissance,
            'sexe': candidat.sexe,
            'ville': candidat.ville,
            'quartier': candidat.quartier,
            'adresse_actuelle': candidat.adresse_actuelle,
            'photo_url': photo_url,
            
            # Parents
            'nom_pere': candidat.nom_pere,
            'tel_pere': candidat.tel_pere,
            'nom_mere': candidat.nom_mere,
            'tel_mere': candidat.tel_mere,
            
            # Localisation
            'region': {
                'id': candidat.region.id,
                'nom': candidat.region.nom
            } if candidat.region else None,
            'departement': {
                'id': candidat.departement.id,
                'nom': candidat.departement.nom
            } if candidat.departement else None,
            
            # Académique
            'bac': {
                'id': candidat.bac.id,
                'libelle': candidat.bac.libelle
            } if candidat.bac else None,
            'serie': {
                'id': candidat.serie.id,
                'libelle': candidat.serie.libelle
            } if candidat.serie else None,
            'mention': {
                'id': candidat.mention.id,
                'libelle': candidat.mention.libelle
            } if candidat.mention else None,
            'filiere': {
                'id': candidat.filiere.id,
                'libelle': candidat.filiere.libelle
            } if candidat.filiere else None,
            'niveau': {
                'id': candidat.niveau.id,
                'libelle': candidat.niveau.libelle
            } if candidat.niveau else None,
            'centre_examen': {
                'id': candidat.centre_examen.id,
                'nom': candidat.centre_examen.nom
            } if candidat.centre_examen else None,
            'centre_depot': {
                'id': candidat.centre_depot.id,
                'nom': candidat.centre_depot.nom
            } if candidat.centre_depot else None,
            
            'etablissement_origine': candidat.etablissement_origine,
            'annee_obtention_diplome': candidat.annee_obtention_diplome,
            
            # Statut
            'statut_dossier': candidat.statut_dossier,
            'motif_rejet': candidat.motif_rejet,
            'date_validation': candidat.date_validation,
            'date_rejet': candidat.date_rejet,
            'valide_par': {
                'nom': candidat.valide_par.get_full_name()
            } if candidat.valide_par else None,
            
            'created_at': candidat.created_at,
            'updated_at': candidat.updated_at
        },
        'dossier': {
            'id': dossier.id,
            'numero_dossier': dossier.numero_dossier,
            'statut': dossier.statut,
            'annee_scolaire': {
                'id': dossier.annee_scolaire.id,
                'libelle': dossier.annee_scolaire.libelle
            } if dossier.annee_scolaire else None,
            'date_soumission': dossier.date_soumission,
            'date_validation': dossier.date_validation,
            'created_at': dossier.created_at,
            'updated_at': dossier.updated_at
        } if dossier else None,
        'documents': documents_data
    }
    
    return Response(response_data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notifications_view(request):
    """Liste des notifications du candidat"""
    try:
        candidat = request.user.candidat
    except Candidat.DoesNotExist:
        return Response({
            'error': 'Profil candidat non trouvé'
        }, status=status.HTTP_404_NOT_FOUND)
    
    notifications = Notification.objects.filter(candidat=candidat).order_by('-created_at')
    
    notifications_data = [{
        'id': notif.id,
        'titre': notif.titre,
        'message': notif.message,
        'type': notif.type,
        'is_read': notif.is_read,
        'action_url': notif.action_url,
        'action_label': notif.action_label,
        'created_at': notif.created_at,
        'read_at': notif.read_at
    } for notif in notifications]
    
    return Response({
        'notifications': notifications_data,
        'total': len(notifications_data),
        'unread': sum(1 for n in notifications_data if not n['is_read'])
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_notification_read(request, notification_id):
    """Marquer une notification comme lue"""
    try:
        candidat = request.user.candidat
        notification = Notification.objects.get(id=notification_id, candidat=candidat)
        
        if not notification.is_read:
            notification.is_read = True
            notification.read_at = timezone.now()
            notification.save()
        
        return Response({
            'message': 'Notification marquée comme lue'
        }, status=status.HTTP_200_OK)
    
    except Notification.DoesNotExist:
        return Response({
            'error': 'Notification non trouvée'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_all_notifications_read(request):
    """Marquer toutes les notifications comme lues"""
    try:
        candidat = request.user.candidat
        updated = Notification.objects.filter(
            candidat=candidat, 
            is_read=False
        ).update(
            is_read=True,
            read_at=timezone.now()
        )
        
        return Response({
            'message': f'{updated} notification(s) marquée(s) comme lue(s)'
        }, status=status.HTTP_200_OK)
    
    except Candidat.DoesNotExist:
        return Response({
            'error': 'Profil candidat non trouvé'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_notification(request, notification_id):
    """Supprimer une notification"""
    try:
        candidat = request.user.candidat
        notification = Notification.objects.get(id=notification_id, candidat=candidat)
        notification.delete()
        
        return Response({
            'message': 'Notification supprimée'
        }, status=status.HTTP_200_OK)
    
    except Notification.DoesNotExist:
        return Response({
            'error': 'Notification non trouvée'
        }, status=status.HTTP_404_NOT_FOUND)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_welcome_notification(request):
    """Crée la notification d'accueil"""
    candidat = request.user.candidat
    create_notification(
        candidat=candidat,
        titre="🎓 Bienvenue dans votre espace notifications !",
        message="""
        🔔 Vous recevrez vos notifications importantes ici :
        
        ✅ Validation de votre dossier
        💰 Confirmation de paiement  
        📄 Statut des pièces justificatives
        📅 Dates des concours       
        """,              
    )
    return Response({"message": "Notification d'accueil créée !"}, status=201)
class ResponsableFiliereViewSet(viewsets.ViewSet):
    """ViewSet pour les responsables de filière"""
    permission_classes = [IsAuthenticated, IsResponsableFiliere]
    
    @action(detail=False, methods=['get'], url_path='dashboard-stats')
    def dashboard_stats(self, request):
        """Statistiques du tableau de bord du RF"""
        try:
            user = request.user
            rf_profile = user.responsable_filiere_profile
            filiere = rf_profile.filiere
            
            print(f"📊 Chargement stats pour {user.email} - Filière: {filiere.libelle}")
            
            # Candidats de la filière
            candidats_filiere = Candidat.objects.filter(filiere=filiere)
            total = candidats_filiere.count()
            
            # Statistiques des dossiers
            valides = candidats_filiere.filter(statut_dossier='valide').count()
            complets = candidats_filiere.filter(statut_dossier='complet').count()
            en_attente = candidats_filiere.filter(statut_dossier='en_attente').count()
            rejetes = candidats_filiere.filter(statut_dossier='rejete').count()
            actifs = candidats_filiere.filter(user__is_active=True).count()
            
            # Calculs
            taux_validation = round((valides / total) * 100, 2) if total > 0 else 0
            taux_inscription = round(((valides + complets) / total) * 100, 2) if total > 0 else 0
            
            stats = {
                'candidats_total': total,
                'candidats_actifs': actifs,
                'candidats_filiere': total,
                'taux_inscription': taux_inscription,
                'dossiers_en_attente': en_attente,
                'dossiers_complets': complets,
                'dossiers_valides': valides,
                'dossiers_rejetes': rejetes,
                'taux_validation': taux_validation,
                'filiere': {
                    'id': filiere.id,
                    'code': filiere.code,
                    'libelle': filiere.libelle,
                    'nom': filiere.libelle,
                    'quota': filiere.quota if hasattr(filiere, 'quota') else None,
                    'capacite': filiere.quota if hasattr(filiere, 'quota') else None,
                }
            }
            
            print(f"✅ Stats: Total={total}, Validés={valides}, Complets={complets}")
            return Response(stats)
            
        except AttributeError as e:
            print(f"❌ Erreur AttributeError: {e}")
            return Response(
                {'error': 'Profil responsable de filière non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"❌ Erreur: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='mes-candidats')
    def mes_candidats(self, request):
        """Liste des candidats de la filière du RF avec filtres"""
        try:
            user = request.user
            rf_profile = user.responsable_filiere_profile
            filiere = rf_profile.filiere
            
            print(f"👥 Chargement candidats pour filière: {filiere.libelle}")
            
            # Filtres
            statut = request.query_params.get('statut', None)
            search = request.query_params.get('search', None)
            
            # Query de base
            candidats = Candidat.objects.filter(filiere=filiere).select_related(
                'user', 'bac', 'serie', 'mention', 'niveau',
                'centre_examen', 'centre_depot'
            )
            
            # Filtrer par statut
            if statut:
                candidats = candidats.filter(statut_dossier=statut)
            
            # Recherche
            if search:
                candidats = candidats.filter(
                    Q(nom__icontains=search) |
                    Q(prenom__icontains=search) |
                    Q(matricule__icontains=search) |
                    Q(email__icontains=search)
                )
            
            # Pagination
            page = int(request.query_params.get('page', 1))
            per_page = int(request.query_params.get('per_page', 20))
            start = (page - 1) * per_page
            end = start + per_page
            
            total = candidats.count()
            candidats_page = candidats.order_by('-created_at')[start:end]
            
            # Construire la réponse manuellement
            results = []
            for candidat in candidats_page:
                results.append({
                    'id': candidat.id,
                    'matricule': candidat.matricule,
                    'nom': candidat.nom,
                    'prenom': candidat.prenom,
                    'email': candidat.email,
                    'telephone': candidat.telephone,
                    'statut_dossier': candidat.statut_dossier,
                    'date_naissance': candidat.date_naissance,
                    'sexe': candidat.sexe,
                    'serie': {
                        'id': candidat.serie.id,
                        'libelle': candidat.serie.libelle
                    } if candidat.serie else None,
                    'mention': {
                        'id': candidat.mention.id,
                        'libelle': candidat.mention.libelle
                    } if candidat.mention else None,
                    'created_at': candidat.created_at,
                    'updated_at': candidat.updated_at,
                })
            
            print(f"✅ {len(results)} candidats trouvés")
            
            return Response({
                'results': results,
                'count': total,
                'page': page,
                'per_page': per_page,
                'total_pages': (total + per_page - 1) // per_page
            })
            
        except Exception as e:
            print(f"❌ Erreur: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='candidat-detail')
    def candidat_detail(self, request, pk=None):
        """Détails complets d'un candidat"""
        try:
            user = request.user
            rf_profile = user.responsable_filiere_profile
            
            # Vérifier que le candidat appartient à la filière du RF
            candidat = Candidat.objects.select_related(
                'user', 'region', 'departement', 'bac', 'serie',
                'mention', 'filiere', 'niveau', 'centre_examen', 'centre_depot'
            ).get(
                id=pk,
                filiere=rf_profile.filiere
            )
            
            # Récupérer les documents
            documents = Document.objects.filter(candidat=candidat)
            
            # Construire la réponse
            data = {
                'id': candidat.id,
                'matricule': candidat.matricule,
                'nom': candidat.nom,
                'prenom': candidat.prenom,
                'email': candidat.email,
                'telephone': candidat.telephone,
                'telephone_secondaire': candidat.telephone_secondaire,
                'date_naissance': candidat.date_naissance,
                'lieu_naissance': candidat.lieu_naissance,
                'sexe': candidat.sexe,
                'ville': candidat.ville,
                'quartier': candidat.quartier,
                'adresse_actuelle': candidat.adresse_actuelle,
                'statut_dossier': candidat.statut_dossier,
                
                # Parents
                'nom_pere': candidat.nom_pere,
                'tel_pere': candidat.tel_pere,
                'nom_mere': candidat.nom_mere,
                'tel_mere': candidat.tel_mere,
                
                # Localisation
                'region': {
                    'id': candidat.region.id,
                    'nom': candidat.region.nom
                } if candidat.region else None,
                'departement': {
                    'id': candidat.departement.id,
                    'nom': candidat.departement.nom
                } if candidat.departement else None,
                
                # Académique
                'bac': {
                    'id': candidat.bac.id,
                    'libelle': candidat.bac.libelle
                } if candidat.bac else None,
                'serie': {
                    'id': candidat.serie.id,
                    'libelle': candidat.serie.libelle
                } if candidat.serie else None,
                'mention': {
                    'id': candidat.mention.id,
                    'libelle': candidat.mention.libelle
                } if candidat.mention else None,
                'filiere': {
                    'id': candidat.filiere.id,
                    'code': candidat.filiere.code,
                    'libelle': candidat.filiere.libelle
                } if candidat.filiere else None,
                'niveau': {
                    'id': candidat.niveau.id,
                    'libelle': candidat.niveau.libelle
                } if candidat.niveau else None,
                'centre_examen': {
                    'id': candidat.centre_examen.id,
                    'nom': candidat.centre_examen.nom
                } if candidat.centre_examen else None,
                'centre_depot': {
                    'id': candidat.centre_depot.id,
                    'nom': candidat.centre_depot.nom
                } if candidat.centre_depot else None,
                'etablissement_origine': candidat.etablissement_origine,
                'annee_obtention_diplome': candidat.annee_obtention_diplome,
                
                # Photo
                'photo_url': request.build_absolute_uri(
                    settings.MEDIA_URL + candidat.photo_path
                ) if candidat.photo_path else None,
                
                # Documents
                'documents': [
                    {
                        'id': doc.id,
                        'type': doc.type_document,
                        'nom': doc.nom_fichier,
                        'url': request.build_absolute_uri(
                            settings.MEDIA_URL + doc.chemin_fichier
                        ),
                        'is_verified': doc.is_verified,
                        'verified_at': doc.verified_at,
                        'commentaire': doc.commentaire_verification
                    }
                    for doc in documents
                ],
                
                'created_at': candidat.created_at,
                'updated_at': candidat.updated_at
            }
            
            return Response(data)
            
        except Candidat.DoesNotExist:
            return Response(
                {'error': 'Candidat non trouvé ou non autorisé'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='profil-filiere')
    def profil_filiere(self, request):
        """Informations détaillées sur la filière du RF"""
        try:
            user = request.user
            rf_profile = user.responsable_filiere_profile
            filiere = rf_profile.filiere
            
            print(f"📋 Chargement profil filière: {filiere.libelle}")
            
            # Candidats de la filière
            candidats_filiere = Candidat.objects.filter(filiere=filiere)
            candidats_valides = candidats_filiere.filter(statut_dossier='valide')
            
            # Répartition par série
            repartition_serie = list(
                candidats_valides.values('serie__libelle')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            
            # Répartition par mention
            repartition_mention = list(
                candidats_valides.values('mention__libelle')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            
            # Évolution mensuelle (6 derniers mois)
            today = timezone.now()
            evolution = []
            for i in range(6):
                mois_date = today - timedelta(days=30*i)
                debut_mois = mois_date.replace(day=1)
                fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                
                count = candidats_valides.filter(
                    date_validation__gte=debut_mois,
                    date_validation__lte=fin_mois
                ).count() if hasattr(Candidat, 'date_validation') else 0
                
                evolution.insert(0, {
                    'mois': debut_mois.strftime('%B'),
                    'total': count
                })
            
            # Âge moyen
            ages = []
            for c in candidats_valides.filter(date_naissance__isnull=False):
                age = (date.today() - c.date_naissance).days / 365.25
                ages.append(age)
            age_moyen = round(sum(ages) / len(ages), 1) if ages else None
            
            # Calcul places restantes et taux de remplissage
            quota = filiere.quota if hasattr(filiere, 'quota') else None
            places_restantes = max(0, quota - candidats_valides.count()) if quota else 0
            taux_remplissage = round((candidats_valides.count() / quota) * 100, 2) if quota and quota > 0 else 0
            
            data = {
                'id': filiere.id,
                'code': filiere.code,
                'libelle': filiere.libelle,
                'description': filiere.description if hasattr(filiere, 'description') else None,
                'quota': quota,
                'places_restantes': places_restantes,
                'taux_remplissage': taux_remplissage,
                
                # Statistiques avancées
                'statistiques': {
                    'total_candidats': candidats_filiere.count(),
                    'candidats_valides': candidats_valides.count(),
                    'candidats_en_attente': candidats_filiere.filter(statut_dossier__in=['en_attente', 'complet']).count(),
                    'candidats_rejetes': candidats_filiere.filter(statut_dossier='rejete').count(),
                    'age_moyen': age_moyen,
                    'repartition_serie': repartition_serie,
                    'repartition_mention': repartition_mention,
                    'evolution_mensuelle': evolution,
                },
                
                # Responsable
                'responsable': {
                    'nom': user.nom,
                    'prenom': user.prenom,
                    'email': user.email,
                    'telephone': rf_profile.telephone
                }
            }
            
            print(f"✅ Profil chargé avec {data['statistiques']['total_candidats']} candidats")
            return Response(data)
            
        except Exception as e:
            print(f"❌ Erreur: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='export-stats')
    def export_stats(self, request):
        """Exporter les statistiques complètes en CSV"""
        try:
            user = request.user
            rf_profile = user.responsable_filiere_profile
            filiere = rf_profile.filiere
            
            # Créer la réponse CSV
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="statistiques_{filiere.code}_{timezone.now().date()}.csv"'
            
            writer = csv.writer(response)
            
            # En-tête principal
            writer.writerow(['STATISTIQUES DÉTAILLÉES'])
            writer.writerow(['Filière', filiere.libelle])
            writer.writerow(['Code', filiere.code])
            writer.writerow(['Date d\'export', timezone.now().strftime('%d/%m/%Y %H:%M')])
            writer.writerow([])
            
            # Section: Vue d'ensemble
            writer.writerow(['=== VUE D\'ENSEMBLE ==='])
            writer.writerow(['Métrique', 'Valeur'])
            
            candidats_filiere = Candidat.objects.filter(filiere=filiere)
            total = candidats_filiere.count()
            valides = candidats_filiere.filter(statut_dossier='valide').count()
            complets = candidats_filiere.filter(statut_dossier='complet').count()
            en_attente = candidats_filiere.filter(statut_dossier='en_attente').count()
            rejetes = candidats_filiere.filter(statut_dossier='rejete').count()
            
            writer.writerow(['Total candidats', total])
            writer.writerow(['Dossiers validés', valides])
            writer.writerow(['Dossiers complets', complets])
            writer.writerow(['Dossiers en attente', en_attente])
            writer.writerow(['Dossiers rejetés', rejetes])
            writer.writerow(['Taux de validation', f'{round((valides/total)*100, 2) if total > 0 else 0}%'])
            writer.writerow([])
            
            # Section: Liste des candidats validés
            writer.writerow(['=== CANDIDATS VALIDÉS ==='])
            writer.writerow([
                'Matricule', 'Nom', 'Prénom', 'Email', 'Téléphone',
                'Sexe', 'Date naissance', 'Série', 'Mention'
            ])
            
            candidats_valides = candidats_filiere.filter(
                statut_dossier='valide'
            ).select_related('serie', 'mention')
            
            for candidat in candidats_valides:
                writer.writerow([
                    candidat.matricule or 'N/A',
                    candidat.nom,
                    candidat.prenom,
                    candidat.email,
                    candidat.telephone or 'N/A',
                    candidat.sexe or 'N/A',
                    candidat.date_naissance.strftime('%d/%m/%Y') if candidat.date_naissance else 'N/A',
                    candidat.serie.libelle if candidat.serie else 'N/A',
                    candidat.mention.libelle if candidat.mention else 'N/A',
                ])
            
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='valider-dossier')
    def valider_dossier(self, request, pk=None):
        """Valider le dossier d'un candidat"""
        print("\n" + "="*80)
        print("🚀 DÉBUT VALIDATION")
        print("="*80)
        
        try:
            print("1️⃣ Récupération user...")
            user = request.user
            print(f"   User: {user.email}")
            
            print("2️⃣ Récupération profil RF...")
            rf_profile = user.responsable_filiere_profile
            print(f"   RF Filière: {rf_profile.filiere.libelle}")
            
            print(f"3️⃣ Recherche candidat ID={pk}...")
            candidat = Candidat.objects.get(
                id=pk,
                filiere=rf_profile.filiere
            )
            print(f"   ✅ Candidat trouvé: {candidat.matricule}")
            print(f"   Statut: {candidat.statut_dossier}")
            
            if candidat.statut_dossier != 'complet': 
                print(f"   ❌ Statut invalide!")
                return Response(
                    {'error': 'Le dossier doit être complet pour être validé'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            print("4️⃣ Mise à jour statut en BD...")
            candidat.statut_dossier = 'valide'
            candidat.date_validation = timezone.now()
            if hasattr(candidat, 'valide_par'):
                candidat.valide_par = user
            candidat.save()
            print("   ✅ Statut sauvegardé en BD")
            
            # ✅ CRÉER LA NOTIFICATION DE VALIDATION
            print("4️⃣bis Création notification...")
            try:
                from candidats.utils.notifications import create_notification
                
                create_notification(
                    candidat=candidat,
                    titre="🎉 Dossier Validé !",
                    message=f"Félicitations ! Votre dossier a été validé le {timezone.now().strftime('%d/%m/%Y à %H:%M')}. Vous pouvez maintenant télécharger votre convocation.",
                    type='validation',
                    action_url='/Mon-dossier',
                    action_label='Voir mon dossier'
                )
                print("   ✅ Notification créée")
            except Exception as notif_error:
                print(f"   ⚠️ Erreur création notification: {notif_error}")
            
            print("5️⃣ DÉBUT GÉNÉRATION EMAIL...")
            try:
                print("   📄 Import pdf_generator...")
                from candidats.utils.pdf_generator import generer_fiche_enrollement
                print("   ✅ Import OK")
                
                print("   📄 Appel generer_fiche_enrollement()...")
                pdf_buffer = generer_fiche_enrollement(candidat)
                print("   ✅ Fonction retournée")
                
                print("   📦 Lecture contenu PDF...")
                pdf_size = len(pdf_buffer.getvalue())
                print(f"   ✅ PDF: {pdf_size} octets")
                
                if pdf_size == 0:
                    print("   ❌ PDF VIDE!")
                    raise ValueError("PDF vide")
                
                print("   📧 Import render_to_string...")
                from django.template.loader import render_to_string
                print("   ✅ Import OK")
                
                print("   📧 Préparation contexte...")
                context = {
                    'candidat': candidat,
                    'filiere': candidat.filiere
                }
                print("   ✅ Contexte OK")
                
                print("   📧 Rendu template...")
                html_message = render_to_string(
                    'emails/validation_enrollement.html',
                    context
                )
                print(f"   ✅ Template rendu: {len(html_message)} caractères")
                
                print("   📧 Import EmailMessage...")
                from django.core.mail import EmailMessage
                from django.conf import settings
                print("   ✅ Import OK")
                
                print("   📧 Création EmailMessage...")
                email = EmailMessage(
                    subject=f'✅ Validation de votre enrôlement - {candidat.filiere.libelle}',
                    body=html_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[candidat.email],
                )
                print("   ✅ Email créé")
                
                print("   📧 Définition content_subtype...")
                email.content_subtype = "html"
                print("   ✅ Content type défini")
                
                print("   📎 Attachement PDF...")
                filename = f'Fiche_Enrollement_{candidat.matricule}.pdf'
                email.attach(filename, pdf_buffer.getvalue(), 'application/pdf')
                print(f"   ✅ PDF attaché: {filename}")
                print(f"   📋 Pièces jointes: {len(email.attachments)}")
                
                print("   📤 Envoi email...")
                result = email.send(fail_silently=False)
                print(f"   ✅ Envoi terminé: result={result}")
                
                if result == 1:
                    print(f"   🎉 EMAIL ENVOYÉ à {candidat.email}")
                else:
                    print(f"   ⚠️ result={result} (attendu 1)")
                
            except Exception as email_error:
                print(f"\n❌ ERREUR DANS BLOC EMAIL:")
                print(f"   Type: {type(email_error).__name__}")
                print(f"   Message: {str(email_error)}")
                import traceback
                traceback.print_exc()
                print("   ⚠️ DOSSIER VALIDÉ mais email NON envoyé")
            
            print("6️⃣ Retour Response...")
            print("="*80 + "\n")
            
            return Response({
                'success': True,
                'message': 'Dossier validé avec succès',
                'candidat': {
                    'id': candidat.id,
                    'matricule': candidat.matricule,
                    'statut': candidat.statut_dossier
                }
            })
            
        except Candidat.DoesNotExist:
            print("❌ Candidat non trouvé")
            return Response(
                {'error': 'Candidat non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"❌ ERREUR GÉNÉRALE: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='rejeter-dossier')
    def rejeter_dossier(self, request, pk=None):
        """Rejeter le dossier d'un candidat avec motif"""
        try:
            user = request.user
            rf_profile = user.responsable_filiere_profile
            motif = request.data.get('motif', '')
            
            if not motif:
                return Response(
                    {'error': 'Le motif de rejet est obligatoire'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            candidat = Candidat.objects.get(
                id=pk,
                filiere=rf_profile.filiere
            )
            
            # Rejeter le dossier
            candidat.statut_dossier = 'rejete'
            if hasattr(candidat, 'motif_rejet'):
                candidat.motif_rejet = motif
            if hasattr(candidat, 'date_rejet'):
                candidat.date_rejet = timezone.now()
            if hasattr(candidat, 'rejete_par'):
                candidat.rejete_par = user
            candidat.save()
            
            # ✅ CRÉER LA NOTIFICATION DE REJET
            print("📢 Création notification de rejet...")
            try:
                from candidats.utils.notifications import create_notification
                
                create_notification(
                    candidat=candidat,
                    titre="❌ Dossier Non Validé",
                    message=f"Votre dossier n'a pas pu être validé. Motif: {motif}. Vous pouvez le corriger et le soumettre à nouveau.",
                    type='rejection',
                    action_url='/Mon-dossier',
                    action_label='Consulter mon dossier'
                )
                print("   ✅ Notification créée")
            except Exception as notif_error:
                print(f"   ⚠️ Erreur création notification: {notif_error}")
            
            # Envoyer l'email
            try:
                from django.template.loader import render_to_string
                from django.core.mail import EmailMessage
                from django.conf import settings
                
                context = {
                    'candidat': candidat,
                    'motif': motif,
                    'filiere': candidat.filiere
                }
                
                html_message = render_to_string(
                    'emails/rejet_enrollement.html',
                    context
                )
                
                email = EmailMessage(
                    subject=f'📋 Notification concernant votre dossier - {candidat.filiere.libelle}',
                    body=html_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[candidat.email],
                )
                
                email.content_subtype = "html"
                email.send(fail_silently=False)
                
                print(f"✅ Email de rejet envoyé à {candidat.email}")
                
            except Exception as email_error:
                print(f"❌ Erreur envoi email: {email_error}")
                import traceback
                traceback.print_exc()
            
            return Response({
                'success': True,
                'message': 'Dossier rejeté avec succès',
                'candidat': {
                    'id': candidat.id,
                    'matricule': candidat.matricule,
                    'statut': candidat.statut_dossier,
                    'motif_rejet': motif
                }
            })
            
        except Candidat.DoesNotExist:
            return Response(
                {'error': 'Candidat non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class AdminAcademiqueViewSet(viewsets.ViewSet):
    """ViewSet pour l'administrateur académique"""
    permission_classes = [IsAuthenticated, IsAdminAcademique]

    @action(detail=False, methods=['get'], url_path='dashboard-stats')
    def dashboard_stats(self, request):
        """Statistiques du dashboard admin académique"""
        try:
            from configurations.models import Filiere
            
            candidats_total = Candidat.objects.count()
            candidats_actifs = Candidat.objects.filter(user__is_active=True).count()
            
            dossiers_valides = Candidat.objects.filter(statut_dossier='valide').count()
            dossiers_complets = Candidat.objects.filter(statut_dossier='complet').count()
            dossiers_en_attente = Candidat.objects.filter(statut_dossier='en_attente').count()
            dossiers_rejetes = Candidat.objects.filter(statut_dossier='rejete').count()
            
            responsables_filieres = User.objects.filter(
                role='responsable_filiere',
                is_active=True
            ).count()
            filieres_actives = Filiere.objects.filter(is_active=True).count()
            
            taux_validation = round(
                (dossiers_valides / candidats_total * 100) if candidats_total > 0 else 0,
                2
            )
            
            semaine_debut = timezone.now() - timedelta(days=7)
            candidats_nouveaux = Candidat.objects.filter(
                created_at__gte=semaine_debut
            ).count()

            taux_rejet = round(
                (dossiers_rejetes / candidats_total * 100) if candidats_total > 0 else 0,
                2
            )
            
            alertes = []
            
            filieres_pleines = Filiere.objects.annotate(
                valides_count=Count('candidats', filter=Q(candidats__statut_dossier='valide'))
            ).filter(
                valides_count__gte=F('quota'),
                is_active=True
            )
            
            for filiere in filieres_pleines:
                alertes.append(
                    f"Filière {filiere.libelle} : quota atteint ({filiere.valides_count}/{filiere.quota})"
                )
            
            old_pending = Candidat.objects.filter(
                statut_dossier='en_attente',
                created_at__lt=timezone.now() - timedelta(days=30)
            ).count()
            
            if old_pending > 0:
                alertes.append(f"{old_pending} dossiers en attente depuis plus de 30 jours")

            stats = {
                'candidats_total': candidats_total,
                'candidats_actifs': candidats_actifs,
                'candidats_nouveaux': candidats_nouveaux, 
                'dossiers_valides': dossiers_valides,
                'dossiers_complets': dossiers_complets,
                'dossiers_en_attente': dossiers_en_attente,
                'dossiers_rejetes': dossiers_rejetes,
                'responsables_filieres': responsables_filieres,
                'filieres_actives': filieres_actives,
                'taux_validation': taux_validation,
                'taux_rejet': taux_rejet,
                'alertes': alertes,
            }
            
            return Response(stats)
            
        except Exception as e:
            print(f"❌ Erreur dashboard stats: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='stats-filieres')
    def stats_filieres(self, request):
        """Statistiques détaillées par filière"""
        try:
            from configurations.models import Filiere
            
            filieres = Filiere.objects.all()
            stats = []
            
            for filiere in filieres:
                candidats = Candidat.objects.filter(filiere=filiere)
                
                responsable_info = None
                try:
                    resp_user = User.objects.filter(
                        role='responsable_filiere',
                        responsable_filiere_profile__filiere=filiere
                    ).first()
                    
                    if resp_user:
                        responsable_info = {
                            'nom': resp_user.nom,
                            'prenom': resp_user.prenom,
                            'email': resp_user.email
                        }
                except Exception:
                    responsable_info = None
                
                stats.append({
                    'id': filiere.id,
                    'code': filiere.code,
                    'libelle': filiere.libelle,
                    'total': candidats.count(),
                    'valides': candidats.filter(statut_dossier='valide').count(),
                    'en_attente': candidats.filter(statut_dossier__in=['en_attente', 'complet']).count(),
                    'rejetes': candidats.filter(statut_dossier='rejete').count(),
                    'quota': getattr(filiere, 'quota', 100),
                    'responsable': responsable_info,
                    'is_active': filiere.is_active,
                })
            
            return Response(stats)
            
        except Exception as e:
            print(f"❌ Erreur stats filières: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='filieres-responsables')
    def filieres_responsables(self, request):
        """Liste des filières avec leurs responsables et statistiques"""
        try:
            from configurations.models import Filiere
            
            filiere_id = request.query_params.get('filiere_id')
            is_active = request.query_params.get('is_active')
            
            filieres_query = Filiere.objects.all()
            
            if filiere_id:
                filieres_query = filieres_query.filter(id=filiere_id)
            if is_active is not None:
                is_active_bool = is_active.lower() == 'true'
                filieres_query = filieres_query.filter(is_active=is_active_bool)
            
            data = []
            
            for filiere in filieres_query:
                responsable_info = None
                try:
                    resp_user = User.objects.filter(
                        role='responsable_filiere',
                        responsable_filiere_profile__filiere=filiere,
                        is_active=True
                    ).first()
                    
                    if resp_user:
                        responsable_info = {
                            'id': resp_user.id,
                            'nom': resp_user.nom,
                            'prenom': resp_user.prenom,
                            'email': resp_user.email,
                            'telephone': getattr(resp_user.responsable_filiere_profile, 'telephone', ''),
                        }
                except Exception as e:
                    print(f"Erreur récupération responsable pour filière {filiere.id}: {e}")
                    responsable_info = None
                
                candidats = Candidat.objects.filter(filiere=filiere)
                total = candidats.count()
                valides = candidats.filter(statut_dossier='valide').count()
                en_attente = candidats.filter(statut_dossier__in=['en_attente', 'complet']).count()
                rejetes = candidats.filter(statut_dossier='rejete').count()
                
                status_val = 'active' if filiere.is_active else 'inactive'
                
                data.append({
                    'id': filiere.id,
                    'code': filiere.code,
                    'libelle': filiere.libelle,
                    'quota': getattr(filiere, 'quota', 100),
                    'status': status_val,
                    'is_active': filiere.is_active,
                    'responsable': responsable_info,
                    'total': total,
                    'valides': valides,
                    'en_attente': en_attente,
                    'rejetes': rejetes,
                    'created_at': filiere.created_at if hasattr(filiere, 'created_at') else None,
                })
            
            return Response(data)
            
        except Exception as e:
            print(f"❌ Erreur filieres_responsables: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='utilisateurs')
    def get_users(self, request):
        """Liste tous les utilisateurs avec filtres"""
        try:
            print("📋 Chargement utilisateurs...")
            
            role = request.query_params.get('role')
            is_active = request.query_params.get('is_active')
            
            users = User.objects.all().order_by('-created_at')
            
            if role:
                users = users.filter(role=role)
            if is_active is not None:
                is_active_bool = is_active.lower() == 'true'
                users = users.filter(is_active=is_active_bool)
            
            data = []
            for user in users:
                data.append({
                    'id': user.id,
                    'nom': user.nom,
                    'prenom': user.prenom,
                    'email': user.email,
                    'role': user.role,
                    'is_active': user.is_active,
                    'created_at': user.created_at.isoformat() if user.created_at else None,
                })
            
            print(f"✅ {len(data)} utilisateurs trouvés")
            return Response(data)
            
        except Exception as e:
            print(f"❌ Erreur get_users: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='export-users')
    def export_users(self, request):
        """Export Excel de tous les utilisateurs"""
        try:
            print("📥 Export utilisateurs Excel demandé")
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Utilisateurs"
            
            # Style pour l'en-tête
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # En-têtes
            headers = ['ID', 'Nom', 'Prénom', 'Email', 'Rôle', 'Actif', 'Date création']
            ws.append(headers)
            
            # Appliquer le style à l'en-tête
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Récupérer et ajouter les données
            users = User.objects.all().order_by('-created_at')
            
            for user in users:
                ws.append([
                    user.id,
                    user.nom,
                    user.prenom,
                    user.email,
                    user.get_role_display(),
                    'Oui' if user.is_active else 'Non',
                    user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else 'N/A'
                ])
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Créer la réponse HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="utilisateurs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Sauvegarder dans la réponse
            wb.save(response)
            
            print(f"✅ Export Excel de {users.count()} utilisateurs")
            return response
            
        except Exception as e:
            print(f"❌ Erreur export_users: {e}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='export-stats')
    def export_stats(self, request):
        """Exporter les statistiques globales en CSV"""
        try:
            from configurations.models import Filiere
            
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="statistiques_globales_{timezone.now().date()}.csv"'
            
            response.write('\ufeff')
            writer = csv.writer(response)
            
            writer.writerow(['STATISTIQUES GLOBALES - ADMIN ACADÉMIQUE'])
            writer.writerow(['Date export', timezone.now().strftime('%d/%m/%Y %H:%M')])
            writer.writerow([])
            
            writer.writerow(['=== VUE ENSEMBLE ==='])
            writer.writerow(['Métrique', 'Valeur'])
            
            candidats_total = Candidat.objects.count()
            writer.writerow(['Total candidats', candidats_total])
            writer.writerow(['Dossiers validés', Candidat.objects.filter(statut_dossier='valide').count()])
            writer.writerow(['Dossiers en attente', Candidat.objects.filter(statut_dossier='en_attente').count()])
            writer.writerow(['Dossiers rejetés', Candidat.objects.filter(statut_dossier='rejete').count()])
            writer.writerow(['Responsables actifs', User.objects.filter(role='responsable_filiere', is_active=True).count()])
            writer.writerow(['Filières actives', Filiere.objects.filter(is_active=True).count()])
            writer.writerow([])
            
            writer.writerow(['=== PAR FILIÈRE ==='])
            writer.writerow(['Filière', 'Code', 'Total', 'Validés', 'En attente', 'Rejetés', 'Quota'])
            
            for filiere in Filiere.objects.all():
                candidats = Candidat.objects.filter(filiere=filiere)
                writer.writerow([
                    filiere.libelle,
                    filiere.code,
                    candidats.count(),
                    candidats.filter(statut_dossier='valide').count(),
                    candidats.filter(statut_dossier='en_attente').count(),
                    candidats.filter(statut_dossier='rejete').count(),
                    getattr(filiere, 'quota', 'N/A')
                ])
            
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )